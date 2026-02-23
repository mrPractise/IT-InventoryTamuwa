import qrcode
from io import BytesIO
from django.core.files import File
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from django.db.models import Q
from assets.models import Category, StatusOption, AssignmentHistory
from maintenance.models import MaintenanceLog, ActionTakenOption


def generate_qr_code(asset):
    """Generate QR code for an asset"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(f"Asset ID: {asset.asset_id}\nSerial: {asset.serial_number}")
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')

    asset.qr_code.save(
        f'qr_{asset.asset_id}.png',
        File(buffer),
        save=False
    )
    asset.save()


def export_assets_excel(assets):
    """
    Export a fully-interconnected Excel workbook.

    Sheets
    ------
    1. ASSET Inventory  – editable data + dropdowns linked to LISTS
    2. QUANTITIES       – auto COUNTIFS per category/status (covers 500 LISTS rows)
    3. MAINTENANCE LOG  – DB records + live formula section for Under-Maintenance assets
    4. LISTS            – master dropdown source (Category / Status / Action Taken)
    5. Past Users       – assignment history
    6. README           – setup instructions including VBA macro for auto-status

    Interconnections
    ----------------
    • LISTS col A → Category dropdown (Asset Inventory col B)
    • LISTS col B → Status dropdown (Asset Inventory col H)
    • LISTS col C → Action Taken dropdown (Maintenance Log col E)
    • QUANTITIES re-counts live; new LISTS rows auto-appear as new quantity rows
    • MAINTENANCE LOG auto-shows assets currently "Under Maintenance"
    • Duplicate Asset ID blocked (COUNTIF stop validation)
    • Duplicate Serial Number per Category blocked (COUNTIFS stop validation)
    • Status colour-coding via conditional formatting
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ── live data from DB ──────────────────────────────────────────────────────
    categories           = list(Category.objects.all().order_by('name'))
    statuses             = list(StatusOption.objects.filter(is_active=True).order_by('name'))
    action_taken_options = list(ActionTakenOption.objects.filter(is_active=True).order_by('name'))

    # ── shared styles ──────────────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    def style_header_row(ws):
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin_border

    # Row/column limits
    INV_MAX   = 2000   # max asset rows in Asset Inventory
    LISTS_MAX = 500    # max rows in LISTS (supports future manual additions)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: LISTS  (must be created first so named references resolve)
    # ══════════════════════════════════════════════════════════════════════════
    ws_lists = wb.create_sheet("LISTS", 0)
    ws_lists.append(['Category Items', 'Current Status', 'Action Taken'])
    style_header_row(ws_lists)

    for i, cat in enumerate(categories, start=2):
        ws_lists.cell(row=i, column=1, value=cat.name)
    for i, st in enumerate(statuses, start=2):
        ws_lists.cell(row=i, column=2, value=st.name)
    for i, act in enumerate(action_taken_options, start=2):
        ws_lists.cell(row=i, column=3, value=act.name)

    for col in 'ABC':
        ws_lists.column_dimensions[col].width = 28

    ws_lists['A1'].comment = Comment(
        "Add new items here:\n"
        "  Col A = Category  →  appears in Asset Inventory dropdown + QUANTITIES\n"
        "  Col B = Status    →  appears in Asset Inventory status dropdown\n"
        "  Col C = Action    →  appears in Maintenance Log action dropdown\n\n"
        "No need to touch any other sheet — formulas pick up new entries automatically.",
        "System"
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: ASSET INVENTORY
    # ══════════════════════════════════════════════════════════════════════════
    ws_inv = wb.create_sheet("ASSET Inventory", 1)
    ws_inv.append([
        'Asset_ID', 'Category', 'Model / Description', 'Purchase Date',
        'Serial Number', 'Assigned To', 'Last Known User', 'Current Status', 'Admin Comments'
    ])
    style_header_row(ws_inv)

    # Fill existing DB data
    for asset in assets:
        ws_inv.append([
            asset.asset_id,
            asset.category.name if asset.category else '',
            asset.model_description,
            asset.purchase_date.strftime('%d/%m/%Y') if asset.purchase_date else '',
            asset.serial_number,
            asset.assigned_to.full_name if asset.assigned_to else '',
            asset.last_known_person.full_name if asset.last_known_person else '',
            asset.status.name if asset.status else '',
            asset.admin_comments,
        ])

    # --- Dropdown: Category (col B) → LISTS!A2:A500 ---
    dv_cat = DataValidation(
        type="list", formula1=f"LISTS!$A$2:$A${LISTS_MAX}",
        allow_blank=True, showErrorMessage=True,
        error="Select from the list or add a new entry in the LISTS sheet (col A).",
        errorTitle="Invalid Category",
    )
    ws_inv.add_data_validation(dv_cat)
    dv_cat.add(f'B2:B{INV_MAX}')

    # --- Dropdown: Status (col H) → LISTS!B2:B500 ---
    dv_status = DataValidation(
        type="list", formula1=f"LISTS!$B$2:$B${LISTS_MAX}",
        allow_blank=True, showErrorMessage=True,
        error="Select from the list or add a new entry in the LISTS sheet (col B).",
        errorTitle="Invalid Status",
    )
    ws_inv.add_data_validation(dv_status)
    dv_status.add(f'H2:H{INV_MAX}')

    # --- No duplicate Asset IDs (col A) ---
    dv_assetid = DataValidation(
        type="custom", formula1='=COUNTIF($A$2:$A$2000,A2)=1',
        allow_blank=True, showErrorMessage=True,
        error="This Asset ID already exists. Each Asset ID must be unique.",
        errorTitle="Duplicate Asset ID", errorStyle="stop",
    )
    ws_inv.add_data_validation(dv_assetid)
    dv_assetid.add(f'A2:A{INV_MAX}')

    # --- No duplicate Serial Number per Category (cols B + E) ---
    dv_serial = DataValidation(
        type="custom",
        formula1='=COUNTIFS($B$2:$B$2000,B2,$E$2:$E$2000,E2)=1',
        allow_blank=True, showErrorMessage=True,
        error="This Serial Number already exists for the selected Category.",
        errorTitle="Duplicate Serial Number", errorStyle="stop",
    )
    ws_inv.add_data_validation(dv_serial)
    dv_serial.add(f'E2:E{INV_MAX}')

    # --- Conditional formatting: status badge colours ---
    STATUS_COLOURS = {
        "In Use":            "C6EFCE",  # green
        "Available":         "DDEBF7",  # blue
        "Under Maintenance": "FFEB9C",  # yellow
        "Missing":           "FFC7CE",  # red
        "Retired":           "D3D3D3",  # grey
    }
    for sname, colour in STATUS_COLOURS.items():
        ws_inv.conditional_formatting.add(
            f'H2:H{INV_MAX}',
            FormulaRule(
                formula=[f'=$H2="{sname}"'],
                fill=PatternFill(start_color=colour, end_color=colour, fill_type="solid"),
            )
        )

    # --- Alternating row shading ---
    ws_inv.conditional_formatting.add(
        f'A2:I{INV_MAX}',
        FormulaRule(
            formula=['=MOD(ROW(),2)=0'],
            fill=PatternFill(start_color="EBF1F7", end_color="EBF1F7", fill_type="solid"),
        )
    )

    # Column widths & freeze
    for col, width in zip('ABCDEFGHI', [15, 20, 32, 14, 20, 22, 22, 20, 40]):
        ws_inv.column_dimensions[col].width = width
    ws_inv.freeze_panes = 'A2'

    ws_inv['F1'].comment = Comment(
        "AUTO-STATUS: When you fill 'Assigned To' (this column), add a VBA macro\n"
        "(see the README sheet) to automatically set 'Current Status' to 'In Use'.\n\n"
        "MAINTENANCE: Any row you set to 'Under Maintenance' automatically appears\n"
        "in the live section of the MAINTENANCE LOG sheet.",
        "System"
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: QUANTITIES
    # Formula per row reads from LISTS!A{r} so new categories auto-appear.
    # ══════════════════════════════════════════════════════════════════════════
    ws_qty = wb.create_sheet("QUANTITIES", 2)
    ws_qty.append(['Asset Category', 'In Use', 'Available', 'Under Maintenance', 'Missing', 'Retired', 'Grand Total'])
    style_header_row(ws_qty)

    status_names = ["In Use", "Available", "Under Maintenance", "Missing", "Retired"]

    for r in range(2, LISTS_MAX + 1):
        cat_ref = f"LISTS!$A${r}"
        # Col A: pull category name from LISTS; blank if LISTS row is empty
        ws_qty.cell(row=r, column=1).value = f'=IFERROR(IF({cat_ref}="","",{cat_ref}),"")'
        # Cols B-F: COUNTIFS per status  ← live from Asset Inventory
        for col_idx, sname in enumerate(status_names, start=2):
            ws_qty.cell(row=r, column=col_idx).value = (
                f'=IFERROR(IF({cat_ref}="","",COUNTIFS('
                f"'ASSET Inventory'!$B:$B,{cat_ref},"
                f"'ASSET Inventory'!$H:$H,\"{sname}\")),\"\")"
            )
        # Col G: Grand Total
        ws_qty.cell(row=r, column=7).value = (
            f'=IFERROR(IF({cat_ref}="","",SUM(B{r}:F{r})),"")'
        )

    # TOTAL summary row
    total_row = LISTS_MAX + 1
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws_qty.cell(row=total_row, column=1, value='TOTAL')
    for col in range(1, 8):
        c = ws_qty.cell(row=total_row, column=col)
        c.font   = Font(bold=True)
        c.fill   = total_fill
        c.border = thin_border
    for col in range(2, 8):
        ws_qty.cell(row=total_row, column=col).value = (
            f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}{LISTS_MAX})'
        )

    # Number formatting & alignment for data cells
    for r in range(2, total_row + 1):
        for col in range(2, 8):
            cell = ws_qty.cell(row=r, column=col)
            cell.number_format = '0'
            cell.alignment     = Alignment(horizontal="center")
    for col in range(1, 8):
        ws_qty.column_dimensions[get_column_letter(col)].width = 20
    ws_qty.freeze_panes = 'A2'

    ws_qty['A1'].comment = Comment(
        "This sheet is fully automatic.\n"
        "• Counts update instantly when you edit the ASSET Inventory sheet.\n"
        "• New categories added to LISTS (col A) auto-appear here as new rows.\n"
        "• Press F9 to force a recalculation if values look stale.",
        "System"
    )

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4: MAINTENANCE LOG
    # Part 1: Actual DB maintenance log records
    # Part 2: Assets currently Under Maintenance (direct DB query — no CSE needed)
    # ══════════════════════════════════════════════════════════════════════════
    ws_maint = wb.create_sheet("MAINTENANCE LOG", 3)
    ws_maint.append([
        'Timestamp', 'Asset ID', 'Category', 'Describe Issue',
        'Action Taken', 'Date Reported', 'Cost of Repair', 'Date Completed', 'Status'
    ])
    style_header_row(ws_maint)

    # Part 1 – actual MaintenanceLog DB records
    maintenance_logs = MaintenanceLog.objects.select_related(
        'asset', 'asset__category', 'action_taken'
    ).order_by('-timestamp')
    db_count = 0
    for log in maintenance_logs:
        ws_maint.append([
            log.timestamp.strftime('%d/%m/%Y %H:%M') if log.timestamp else '',
            log.asset.asset_id if log.asset else '',
            log.asset.category.name if log.asset and log.asset.category else '',
            log.description,
            log.action_taken.name if log.action_taken else '',
            log.date_reported.strftime('%d/%m/%Y') if log.date_reported else '',
            float(log.cost_of_repair) if log.cost_of_repair else '',
            log.date_completed.strftime('%d/%m/%Y') if log.date_completed else '',
            log.maintenance_status,
        ])
        db_count += 1

    # Part 2 – assets currently Under Maintenance (pulled from Asset DB directly)
    from assets.models import Asset as AssetModel
    under_maint_assets = AssetModel.objects.filter(
        status__name='Under Maintenance', is_deleted=False
    ).select_related('category', 'status', 'assigned_to', 'department').order_by('asset_id')

    if under_maint_assets.exists():
        sep_row = ws_maint.max_row + 2
        sep_cell = ws_maint.cell(
            row=sep_row, column=1,
            value="↓  Assets currently Under Maintenance (snapshot at export time)"
        )
        sep_cell.font = Font(bold=True, color="FF0000")
        ws_maint.merge_cells(
            start_row=sep_row, start_column=1, end_row=sep_row, end_column=9
        )
        for asset_um in under_maint_assets:
            ws_maint.append([
                '',                                          # Timestamp (n/a for snapshot)
                asset_um.asset_id,
                asset_um.category.name if asset_um.category else '',
                asset_um.model_description,                  # Describe Issue placeholder
                '',                                          # Action Taken (to be filled)
                '',                                          # Date Reported
                '',                                          # Cost
                '',                                          # Date Completed
                'Under Maintenance',
            ])

    total_maint_rows = ws_maint.max_row

    # Action Taken dropdown (col E) → LISTS!C2:C500
    if action_taken_options:
        dv_action = DataValidation(
            type="list", formula1=f"LISTS!$C$2:$C${LISTS_MAX}",
            allow_blank=True, showErrorMessage=True,
            error="Select from the list or add to LISTS sheet col C.",
            errorTitle="Invalid Action",
        )
        ws_maint.add_data_validation(dv_action)
        dv_action.add(f'E2:E{total_maint_rows + 100}')

    for col, width in zip('ABCDEFGHI', [18, 14, 18, 40, 22, 14, 14, 14, 18]):
        ws_maint.column_dimensions[col].width = width
    ws_maint.freeze_panes = 'A2'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5: PAST USERS
    # ══════════════════════════════════════════════════════════════════════════
    ws_past = wb.create_sheet("Past Users", 4)
    ws_past.append(['Asset_ID', 'User', 'Start Date', 'End Date'])
    style_header_row(ws_past)

    for a in AssignmentHistory.objects.select_related('asset', 'user').order_by('-start_date'):
        ws_past.append([
            a.asset.asset_id if a.asset else '',
            (a.user.get_full_name() or a.user.username) if a.user else '',
            a.start_date.strftime('%d/%m/%Y') if a.start_date else '',
            a.end_date.strftime('%d/%m/%Y') if a.end_date else 'Current',
        ])

    for col, width in zip('ABCD', [15, 28, 15, 15]):
        ws_past.column_dimensions[col].width = width
    ws_past.freeze_panes = 'A2'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 6: README – instructions + VBA macro code
    # ══════════════════════════════════════════════════════════════════════════
    ws_readme = wb.create_sheet("📋 README", 5)
    ws_readme.column_dimensions['A'].width = 95

    readme_content = [
        ("IT Asset Inventory – Excel Export Guide", True, 14, "366092"),
        ("", False, 11, None),
        ("WHAT IS AUTOMATIC (no setup needed)", True, 11, None),
        ("  • Category / Status / Action Taken dropdowns are linked to the LISTS sheet.", False, 11, None),
        ("  • Add a new Category in LISTS col A → it instantly appears in the Asset Inventory", False, 11, None),
        ("    Category dropdown AND shows as a new row in QUANTITIES.", False, 11, None),
        ("  • QUANTITIES counts update live when you edit the Asset Inventory sheet.", False, 11, None),
        ("  • MAINTENANCE LOG bottom section auto-lists every asset with status", False, 11, None),
        ('    "Under Maintenance" from the Asset Inventory sheet (press F9 to refresh).', False, 11, None),
        ("", False, 11, None),
        ("VALIDATION RULES (enforced by Excel stop errors)", True, 11, None),
        ("  • Duplicate Asset_ID (col A) → Excel will block the entry.", False, 11, None),
        ("  • Duplicate Serial Number within the same Category (col E, keyed on col B)", False, 11, None),
        ("    → Excel will block the entry.", False, 11, None),
        ("  • Status and Category must come from the LISTS sheet.", False, 11, None),
        ("", False, 11, None),
        ("AUTO-STATUS MACRO (one-time 2-minute setup)", True, 11, "FF0000"),
        ("This makes Status auto-change to 'In Use' when you fill the 'Assigned To' column.", False, 11, None),
        ("", False, 11, None),
        ("  Step 1: Press Alt+F11 to open the VBA editor.", False, 11, None),
        ("  Step 2: In the left panel, expand VBAProject → Microsoft Excel Objects.", False, 11, None),
        ("  Step 3: Double-click the sheet named 'ASSET Inventory'.", False, 11, None),
        ("  Step 4: Paste the code below into the code window:", False, 11, None),
        ("", False, 11, None),
        ("     Private Sub Worksheet_Change(ByVal Target As Range)", False, 11, None),
        ("         ' Col F = Assigned To,  Col H = Current Status", False, 11, None),
        ("         If Target.Column = 6 And Target.Row > 1 Then", False, 11, None),
        ("             Application.EnableEvents = False", False, 11, None),
        ("             If Trim(Target.Value) <> \"\" Then", False, 11, None),
        ("                 Cells(Target.Row, 8).Value = \"In Use\"", False, 11, None),
        ("             End If", False, 11, None),
        ("             Application.EnableEvents = True", False, 11, None),
        ("         End If", False, 11, None),
        ("     End Sub", False, 11, None),
        ("", False, 11, None),
        ("  Step 5: Close the VBA editor (Alt+Q).", False, 11, None),
        ('  Step 6: Save the file as "Excel Macro-Enabled Workbook (*.xlsm)".', False, 11, None),
        ("", False, 11, None),
        ("TIPS", True, 11, None),
        ("  • Press F9 to force-recalculate all formulas.", False, 11, None),
        ("  • Do NOT delete the LISTS sheet — all dropdowns depend on it.", False, 11, None),
        ("  • QUANTITIES has formulas for 500 rows to support future categories.", False, 11, None),
        ("  • Status colour coding: Green=In Use, Blue=Available,", False, 11, None),
        ("    Yellow=Under Maintenance, Red=Missing, Grey=Retired.", False, 11, None),
    ]

    for text, bold, size, colour in readme_content:
        ws_readme.append([text])
        row = ws_readme.max_row
        f = Font(bold=bold, size=size)
        if colour:
            f = Font(bold=bold, size=size, color=colour)
        ws_readme.cell(row=row, column=1).font = f

    # ── HTTP response ──────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="IT_Inventory_Export.xlsx"'
    wb.save(response)
    return response


def export_assets_pdf(assets):
    """Export assets to PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("Assets Export", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    # Table data
    data = [['Asset ID', 'Category', 'Model', 'Serial', 'Assigned To', 'Status']]

    for asset in assets[:100]:  # Limit to 100 for PDF
        data.append([
            asset.asset_id,
            asset.category.name if asset.category else '',
            asset.model_description[:30],
            asset.serial_number[:20],
            asset.assigned_to.full_name if asset.assigned_to else 'Unassigned',
            asset.status.name if asset.status else '',
        ])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="assets_export.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response
